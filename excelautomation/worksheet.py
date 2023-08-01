
'''
Created on 01.08.2023

@author: neumann
'''

import openpyxl
from pprint import pprint
from protocol_reader import open_protocol

def print_range(worksheet, bereich):
    rows_start = int(bereich.split(':')[0][1:])
    rows_end = int(bereich.split(':')[1][1:])
    column_start = bereich.split(':')[0][0]
    column_end = bereich.split(':')[1][0]
    
    for z in range(rows_start, rows_end + 1):
        for s in ['A', 'B', 'C']:
            print(str(worksheet[s + str(z)].value), end="\t")
            #print(s + str(z), end="\t")
        print()
    

wb = openpyxl.Workbook()
ws = wb[wb.sheetnames[0]]

protokoll = open_protocol('protocol.txt')

for zeile in protokoll:
    ws.append(zeile[:-1])
    #print(zeile)

ws.auto_filter.ref=f'A1:{openpyxl.utils.get_column_letter(ws.max_column)}{ws.max_row}'

print(f'A1:{openpyxl.utils.get_column_letter(ws.max_column)}{ws.max_row}')

wb.save('protocol.xlsx')




